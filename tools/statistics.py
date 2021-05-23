from openpyxl import Workbook, load_workbook


class Statistics:

    name = "WorldCup"
    teamA = "Hamas"
    teamB = "BeitarYerushlaim"
    Location = "Yarden"
    Date = "23/05/2021"
    Weather = "Rain"

    TeamA_Accuracy_BallIn = 10
    TeamB_Accuracy_BallIn = 90

    TeamA_Accuracy_BallOut = 90
    TeamB_Accuracy_BallOut = 10

    def save(self):
        self.saveStat(self.name, self.teamA, self.teamB, self.Location, self.Date,
                      self.Weather, self.TeamA_Accuracy_BallIn, self.TeamB_Accuracy_BallIn,
                      self.TeamA_Accuracy_BallOut, self.TeamB_Accuracy_BallOut)

    def saveStat(self, name, teamA, teamB, Location, Date, Weather,
                 TeamA_Accuracy_BallIn, TeamB_Accuracy_BallIn,
                 TeamA_Accuracy_BallOut, TeamB_Accuracy_BallOut):

        wb = load_workbook('/home/chameleonvision/Desktop/FinalProject/ChameleonVISION/statistics/statistics.xlsx')
        ws = wb.active
        ws['A2'].value = name
        ws['B2'].value = teamA
        ws['C2'].value = teamB
        ws['D2'].value = Location
        ws['E2'].value = Date
        ws['F2'].value = Weather
        ws['G2'].value = TeamA_Accuracy_BallIn
        ws['H2'].value = TeamB_Accuracy_BallIn
        ws['I2'].value = TeamA_Accuracy_BallOut
        ws['J2'].value = TeamB_Accuracy_BallOut

        wb.save('/home/chameleonvision/Desktop/FinalProject/ChameleonVISION/statistics/statistics.xlsx')




